"""PCBS Excel (.xlsx) parser.

Handles multi-sheet PCBS workbooks where each sheet is a year.
Supports Arabic month names, average columns, and multi-row headers.
"""

from __future__ import annotations

import re
import logging
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)

# Arabic month name → month number
ARABIC_MONTHS: dict[str, int] = {
    # With definite article
    "كانون الثاني": 1,
    "شباط": 2,
    "آذار": 3,
    "نيسان": 4,
    "أيار": 5,
    "حزيران": 6,
    "تموز": 7,
    "آب": 8,
    "أيلول": 9,
    "تشرين الأول": 10,
    "تشرين الثاني": 11,
    "كانون الأول": 12,
    # Without definite article (common in PCBS)
    "كانون ثاني": 1,
    "تشرين أول": 10,
    "تشرين ثاني": 11,
    "كانون أول": 12,
}

# Average keyword in Arabic
AVERAGE_KEYWORDS = ["متوسط", "معدل"]

# Percent change keywords in Arabic
PCT_CHANGE_KEYWORDS = ["نسبة التغير", "نسبة تغير", "% التغير", "التغير %"]

# Known PCBS indicator translations
INDICATOR_TRANSLATIONS: dict[str, str] = {
    "التعدين واستغلال المحاجر": "Mining and Quarrying",
    "الصناعة التحويلية": "Manufacturing",
    "إمدادات الكهرباء والغاز والبخار وتكييف الهواء": "Electricity, Gas, Steam and Air Conditioning Supply",
    "امدادات المياه وانشطة الصرف الصحي  وادارة النفايات ومعالجتها": "Water Supply, Sewerage, Waste Management and Remediation",
    "الرقم القياسي العام لكميات الإنتاج الصناعي": "General Industrial Production Index",
    "الرقم القياسي العام": "General Index",
    "المواد الخام": "Raw Materials",
    "أجور ومصاريف القوى العاملة": "Labour Cost and Wages",
    "استئجار معدات وآليات": "Hiring of Equipment",
    "المؤشر العام": "General Indicator",
    "الأغذية والمشروبات غير الكحولية": "Food and Non-Alcoholic Beverages",
    "المشروبات الكحولية والتبغ": "Alcoholic Beverages and Tobacco",
    "الملابس والأحذية": "Clothing and Footwear",
    "السكن والمياه والكهرباء والغاز وأنواع الوقود": "Housing, Water, Electricity, Gas and Other Fuels",
    "المفروشات والأثاث ولوازم الصيانة المنزلية": "Furnishings, Household Equipment and Maintenance",
    "الصحة": "Health",
    "النقل والمواصلات": "Transport",
    "الاتصالات": "Communication",
    "الترويح والثقافة": "Recreation and Culture",
    "التعليم": "Education",
    "المطاعم والفنادق": "Restaurants and Hotels",
    "سلع وخدمات متنوعة": "Miscellaneous Goods and Services",
}


@dataclass
class ParsedObservation:
    """A single parsed observation from an Excel file."""
    indicator_name_ar: str
    indicator_name_en: str
    time_period: date
    time_precision: str  # "month", "year"
    value: float
    sheet_name: str = ""
    row_num: int = 0
    col_num: int = 0


@dataclass
class ParsedSheet:
    """Results from parsing one Excel file."""
    title_ar: str = ""
    title_en: str = ""
    base_year_note: str = ""
    observations: list[ParsedObservation] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def translate_indicator(name_ar: str) -> str:
    """Translate Arabic indicator name to English, or transliterate."""
    name_ar = name_ar.strip()
    if name_ar in INDICATOR_TRANSLATIONS:
        return INDICATOR_TRANSLATIONS[name_ar]
    # Try partial match
    for ar, en in INDICATOR_TRANSLATIONS.items():
        if ar in name_ar:
            return en
    # Return Arabic as-is if no translation found
    return name_ar


def parse_column_header(header: str) -> tuple[Optional[int], Optional[int], bool, bool]:
    """Parse Arabic column header to extract month and year.

    Returns (month, year, is_average, is_pct_change).
    """
    if not header or not isinstance(header, str):
        return None, None, False, False

    header = header.strip()

    # Check for % change
    is_pct = any(kw in header for kw in PCT_CHANGE_KEYWORDS)
    if is_pct:
        year_match = re.search(r'(\d{4})', header)
        year = int(year_match.group(1)) if year_match else None
        return None, year, False, True

    # Check for average
    is_avg = any(kw in header for kw in AVERAGE_KEYWORDS)

    # Extract year
    year_match = re.search(r'(\d{4})', header)
    year = int(year_match.group(1)) if year_match else None

    if is_avg:
        return None, year, True, False

    # Extract month from Arabic month name
    for ar_month, month_num in sorted(ARABIC_MONTHS.items(), key=lambda x: -len(x[0])):
        if ar_month in header:
            return month_num, year, False, False

    return None, year, False, False


def parse_xlsx(filepath: str | Path) -> ParsedSheet:
    """Parse a PCBS Excel file with year-per-sheet layout."""
    filepath = Path(filepath)
    result = ParsedSheet()

    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    except Exception as e:
        result.errors.append(f"Failed to open {filepath}: {e}")
        return result

    for sheet_name in wb.sheetnames:
        # Check if sheet name is a year
        try:
            sheet_year = int(sheet_name.strip())
        except ValueError:
            logger.debug(f"Skipping non-year sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            result.errors.append(f"Sheet {sheet_name} has too few rows ({len(rows)})")
            continue

        # Row 0: Title (Arabic)
        if not result.title_ar and rows[0][0]:
            result.title_ar = str(rows[0][0]).strip()

        # Row 1: Base year note
        if not result.base_year_note and rows[1][0]:
            result.base_year_note = str(rows[1][0]).strip()

        # Row 2: Column headers
        headers = [str(c).strip() if c else "" for c in rows[2]]

        # Parse each column header
        col_info: list[tuple[Optional[int], Optional[int], bool, bool]] = []
        for h in headers:
            col_info.append(parse_column_header(h))

        # Rows 3+: Data rows
        for row_idx in range(3, len(rows)):
            row = rows[row_idx]
            if not row or not row[0]:
                continue

            indicator_ar = str(row[0]).strip()

            # Skip source/footnote rows
            if indicator_ar.startswith("المصدر") or indicator_ar.startswith("ملاحظة"):
                continue
            if not indicator_ar or indicator_ar == "None":
                continue

            indicator_en = translate_indicator(indicator_ar)

            for col_idx in range(1, min(len(row), len(col_info))):
                value = row[col_idx]
                if value is None:
                    continue
                # Skip formula strings that weren't cached
                if isinstance(value, str) and value.startswith("="):
                    continue
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    continue

                month, year, is_avg, is_pct = col_info[col_idx]

                if is_pct:
                    # % Change column — store as separate indicator
                    pct_name_en = f"{indicator_en} (% Change)"
                    pct_name_ar = f"{indicator_ar} (نسبة التغير)"
                    # Use the sheet year for the % change
                    tp = date(sheet_year, 1, 1)
                    result.observations.append(ParsedObservation(
                        indicator_name_ar=pct_name_ar,
                        indicator_name_en=pct_name_en,
                        time_period=tp,
                        time_precision="year",
                        value=value,
                        sheet_name=sheet_name,
                        row_num=row_idx + 1,
                        col_num=col_idx + 1,
                    ))
                elif is_avg:
                    # Average column — use the year, precision = "year"
                    if year:
                        tp = date(year, 1, 1)
                        result.observations.append(ParsedObservation(
                            indicator_name_ar=indicator_ar,
                            indicator_name_en=indicator_en,
                            time_period=tp,
                            time_precision="year",
                            value=value,
                            sheet_name=sheet_name,
                            row_num=row_idx + 1,
                            col_num=col_idx + 1,
                        ))
                elif month and year:
                    tp = date(year, month, 1)
                    result.observations.append(ParsedObservation(
                        indicator_name_ar=indicator_ar,
                        indicator_name_en=indicator_en,
                        time_period=tp,
                        time_precision="month",
                        value=value,
                        sheet_name=sheet_name,
                        row_num=row_idx + 1,
                        col_num=col_idx + 1,
                    ))
                elif year:
                    # Year-only column
                    tp = date(year, 1, 1)
                    result.observations.append(ParsedObservation(
                        indicator_name_ar=indicator_ar,
                        indicator_name_en=indicator_en,
                        time_period=tp,
                        time_precision="year",
                        value=value,
                        sheet_name=sheet_name,
                        row_num=row_idx + 1,
                        col_num=col_idx + 1,
                    ))

    wb.close()

    # Generate English title from Arabic
    if result.title_ar:
        result.title_en = _translate_title(result.title_ar)

    return result


def _translate_title(title_ar: str) -> str:
    """Generate an English title from the Arabic title."""
    # Common PCBS title patterns
    if "الإنتاج الصناعي" in title_ar or "الانتاج الصناعي" in title_ar:
        return "Monthly Industrial Production Index Numbers by Major Groups"
    if "أسعار تكاليف البناء" in title_ar:
        return "Construction Cost Indices by Major Groups"
    if "الرقم القياسي لأسعار المستهلك" in title_ar:
        return "Consumer Price Index"
    if "الرقم القياسي لأسعار المنتج" in title_ar:
        return "Producer Price Index"
    # Fallback
    return title_ar


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python xlsx_parser.py <file.xlsx>")
        sys.exit(1)

    logging.basicConfig(level=logging.INFO)
    result = parse_xlsx(sys.argv[1])

    print(f"Title (AR): {result.title_ar}")
    print(f"Title (EN): {result.title_en}")
    print(f"Base year: {result.base_year_note}")
    print(f"Observations: {len(result.observations)}")
    print(f"Errors: {result.errors}")

    if result.observations:
        # Group by indicator
        indicators = {}
        for obs in result.observations:
            key = obs.indicator_name_en
            if key not in indicators:
                indicators[key] = []
            indicators[key].append(obs)

        print(f"\nIndicators ({len(indicators)}):")
        for name, obs_list in sorted(indicators.items()):
            monthly = sum(1 for o in obs_list if o.time_precision == "month")
            yearly = sum(1 for o in obs_list if o.time_precision == "year")
            print(f"  {name}: {monthly} monthly + {yearly} yearly = {len(obs_list)} observations")

        # Spot checks
        print("\n=== Spot Checks ===")
        for obs in result.observations:
            if obs.indicator_name_en == "General Industrial Production Index":
                if obs.time_period == date(2011, 1, 1) and obs.time_precision == "month":
                    expected = 90.57
                    match = abs(obs.value - expected) < 0.01
                    print(f"  General Index, Jan 2011: {obs.value:.2f} (expected {expected}) {'✓' if match else '✗'}")
                if obs.time_period == date(2025, 1, 1) and obs.time_precision == "month":
                    expected = 79.25
                    match = abs(obs.value - expected) < 0.01
                    print(f"  General Index, Jan 2025: {obs.value:.2f} (expected {expected}) {'✓' if match else '✗'}")
            if obs.indicator_name_en == "Manufacturing":
                if obs.time_period == date(2020, 6, 1) and obs.time_precision == "month":
                    expected = 89.47
                    match = abs(obs.value - expected) < 0.01
                    print(f"  Manufacturing, Jun 2020: {obs.value:.2f} (expected {expected}) {'✓' if match else '✗'}")
            if obs.indicator_name_en == "General Industrial Production Index":
                if obs.time_period == date(2025, 1, 1) and obs.time_precision == "year":
                    expected = 82.72
                    match = abs(obs.value - expected) < 0.01
                    print(f"  General Index, Avg 2025: {obs.value:.2f} (expected {expected}) {'✓' if match else '✗'}")
            if obs.indicator_name_en == "General Industrial Production Index (% Change)":
                if obs.time_period == date(2025, 1, 1) and obs.time_precision == "year":
                    expected = 3.39
                    match = abs(obs.value - expected) < 0.01
                    print(f"  General Index, % Change 2025: {obs.value:.2f} (expected {expected}) {'✓' if match else '✗'}")
